#!/usr/bin/env python3
"""
Download every worksheet in a public Google Sheets workbook as a separate CSV.

Example:
    python download_google_sheets_csv.py \
        "https://docs.google.com/spreadsheets/d/1dpydoJvidsMgXjAC2tYVspX3I_8icSJ3OKPgP7GhLlc/edit?gid=1853060031"

Dependencies:
    python -m pip install openpyxl
"""

from __future__ import annotations

import argparse
import csv
import re
import shutil
import sys
import tempfile
import urllib.error
import urllib.request
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Iterable, Sequence


SHEET_ID_RE = re.compile(r"/spreadsheets/d/([A-Za-z0-9_-]+)")
INVALID_FILENAME_CHARS_RE = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
WINDOWS_RESERVED_NAMES = {
    "CON", "PRN", "AUX", "NUL",
    *(f"COM{i}" for i in range(1, 10)),
    *(f"LPT{i}" for i in range(1, 10)),
}


def parse_spreadsheet_id(value: str) -> str:
    """Extract a Google spreadsheet ID from a URL, or accept a bare ID."""
    value = value.strip()

    match = SHEET_ID_RE.search(value)
    if match:
        return match.group(1)

    if re.fullmatch(r"[A-Za-z0-9_-]+", value):
        return value

    raise ValueError(
        "Could not extract a spreadsheet ID. Pass a Google Sheets URL "
        "or the bare spreadsheet ID."
    )


def safe_filename(sheet_name: str) -> str:
    """Convert a worksheet title into a safe cross-platform filename stem."""
    name = INVALID_FILENAME_CHARS_RE.sub("_", sheet_name).strip(" .")
    name = re.sub(r"\s+", " ", name)

    if not name:
        name = "sheet"

    if name.upper() in WINDOWS_RESERVED_NAMES:
        name = f"_{name}"

    # Leave room for suffixes and ".csv".
    return name[:180]


def unique_csv_path(
    output_dir: Path,
    sheet_name: str,
    used_names: set[str],
) -> Path:
    """Return a collision-free CSV path for a worksheet."""
    stem = safe_filename(sheet_name)
    candidate = stem
    counter = 2

    while candidate.casefold() in used_names:
        candidate = f"{stem}_{counter}"
        counter += 1

    used_names.add(candidate.casefold())
    return output_dir / f"{candidate}.csv"


def download_workbook(spreadsheet_id: str, destination: Path) -> None:
    """Download a public Google Sheets workbook in XLSX format."""
    url = (
        f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"
        "/export?format=xlsx"
    )
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (X11; Linux x86_64) "
                "AppleWebKit/537.36 Chrome/151 Safari/537.36"
            )
        },
    )

    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            with destination.open("wb") as output:
                shutil.copyfileobj(response, output)
    except urllib.error.HTTPError as exc:
        raise RuntimeError(
            f"Google returned HTTP {exc.code}. Confirm that the spreadsheet "
            "is public and that downloading is permitted."
        ) from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"Download failed: {exc.reason}") from exc

    # XLSX files are ZIP archives. This catches login/error HTML responses.
    with destination.open("rb") as handle:
        signature = handle.read(4)

    if signature != b"PK\x03\x04":
        raise RuntimeError(
            "Google did not return an XLSX workbook. The spreadsheet may not "
            "be public, or download/copy access may be disabled."
        )


def csv_value(value: Any) -> Any:
    """Normalize Excel values for predictable CSV output."""
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.isoformat(sep=" ")

    if isinstance(value, (date, time)):
        return value.isoformat()

    if isinstance(value, timedelta):
        return str(value)

    return value


def trim_trailing_empty_cells(row: Sequence[Any]) -> list[Any]:
    """Remove empty cells at the right edge of a row."""
    normalized = [csv_value(value) for value in row]

    while normalized and normalized[-1] == "":
        normalized.pop()

    return normalized


def write_worksheet_csv(
    worksheet: Any,
    destination: Path,
    keep_trailing_empty_rows: bool = False,
) -> tuple[int, int]:
    """
    Write one openpyxl worksheet to CSV.

    Internal blank rows are retained. Trailing blank rows are removed by
    default to avoid enormous CSV files caused by formatting-only cells.
    """
    written_rows = 0
    widest_row = 0
    pending_blank_rows = 0

    with destination.open("w", encoding="utf-8", newline="") as output:
        writer = csv.writer(output, lineterminator="\n")

        for source_row in worksheet.iter_rows(values_only=True):
            row = trim_trailing_empty_cells(source_row)

            if not row:
                pending_blank_rows += 1
                continue

            for _ in range(pending_blank_rows):
                writer.writerow([])
                written_rows += 1
            pending_blank_rows = 0

            writer.writerow(row)
            written_rows += 1
            widest_row = max(widest_row, len(row))

        if keep_trailing_empty_rows:
            for _ in range(pending_blank_rows):
                writer.writerow([])
                written_rows += 1

    return written_rows, widest_row


def export_all_sheets(
    workbook_path: Path,
    output_dir: Path,
    overwrite: bool,
    keep_trailing_empty_rows: bool,
) -> list[tuple[str, Path, int, int]]:
    """Export every worksheet in an XLSX workbook to a separate CSV."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "The 'openpyxl' package is required. Install it with:\n"
            "  python -m pip install openpyxl\n"
            "or on Arch Linux:\n"
            "  sudo pacman -S python-openpyxl"
        ) from exc

    output_dir.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=True,
    )

    results: list[tuple[str, Path, int, int]] = []
    used_names: set[str] = set()

    try:
        for worksheet in workbook.worksheets:
            destination = unique_csv_path(
                output_dir,
                worksheet.title,
                used_names,
            )

            if destination.exists() and not overwrite:
                raise FileExistsError(
                    f"Refusing to overwrite existing file: {destination}\n"
                    "Run again with --overwrite to replace existing CSV files."
                )

            rows, columns = write_worksheet_csv(
                worksheet,
                destination,
                keep_trailing_empty_rows=keep_trailing_empty_rows,
            )
            results.append(
                (worksheet.title, destination, rows, columns)
            )
    finally:
        workbook.close()

    return results


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Download every worksheet in a public Google Sheets workbook "
            "as a separate CSV file."
        )
    )
    parser.add_argument(
        "sheet",
        help="Public Google Sheets URL or bare spreadsheet ID.",
    )
    parser.add_argument(
        "-o",
        "--output-dir",
        type=Path,
        help=(
            "Directory for CSV files. Default: "
            "<spreadsheet-id>_csv in the current directory."
        ),
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Replace CSV files that already exist.",
    )
    parser.add_argument(
        "--keep-xlsx",
        type=Path,
        metavar="PATH",
        help="Also retain the downloaded XLSX workbook at PATH.",
    )
    parser.add_argument(
        "--keep-trailing-empty-rows",
        action="store_true",
        help=(
            "Retain blank rows after the last nonempty row. Normally these "
            "are removed because formatting can make worksheets appear huge."
        ),
    )
    return parser


def main() -> int:
    parser = build_argument_parser()
    args = parser.parse_args()

    try:
        spreadsheet_id = parse_spreadsheet_id(args.sheet)
        output_dir = args.output_dir or Path(f"{spreadsheet_id}_csv")

        with tempfile.TemporaryDirectory(prefix="google-sheets-") as temp_dir:
            workbook_path = Path(temp_dir) / "workbook.xlsx"
            print(f"Downloading workbook {spreadsheet_id} ...")
            download_workbook(spreadsheet_id, workbook_path)

            if args.keep_xlsx:
                keep_path = args.keep_xlsx.expanduser().resolve()
                keep_path.parent.mkdir(parents=True, exist_ok=True)
                if keep_path.exists() and not args.overwrite:
                    raise FileExistsError(
                        f"Refusing to overwrite existing file: {keep_path}\n"
                        "Run again with --overwrite to replace it."
                    )
                shutil.copy2(workbook_path, keep_path)
                print(f"Saved workbook: {keep_path}")

            results = export_all_sheets(
                workbook_path=workbook_path,
                output_dir=output_dir,
                overwrite=args.overwrite,
                keep_trailing_empty_rows=args.keep_trailing_empty_rows,
            )

        print(f"\nExported {len(results)} worksheet(s) to {output_dir}:")
        for title, path, rows, columns in results:
            print(
                f"  {title!r} -> {path.name} "
                f"({rows} row(s), up to {columns} column(s))"
            )

        return 0

    except (ValueError, RuntimeError, FileExistsError, OSError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
